from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl

chrome_options = webdriver.chrome.options.Options()
chrome_options.add_argument('--headless')
# global chrome_options

# Excel Headings layout:
"Artist	Release Title	LP or EP	Genre	Date	Condition	Discogs Price (£)"

links = """
https://www.discogs.com/Zomby-Mush-Spliff-Dub-Rustie-Remix/release/1827863
https://www.discogs.com/Harry-Craze-Wa6-Thirty-Six-Hours/release/2015339
https://www.discogs.com/Skream-Midnight-Request-Line-I/release/561478
https://www.discogs.com/Joker-City-Hopper/release/2045542
https://www.discogs.com/Terror-Danjah-Acid-ProPlus/release/2132871
https://www.discogs.com/Pinch-2-Ft-Yolanda-Get-Up/release/2054095
https://www.discogs.com/Flying-Lotus-Joker-Glendale-Galleria-Untitled_Rsn/release/1763987
https://www.discogs.com/Various-Hyperdub-52/release/1923860
https://www.discogs.com/The-Bug-Feat-KillaP-Flow-Dan-Skeng/release/1067932
""".split('\n')[1:-1]


def get_info_from_soup(soup):
    data = ['']
    
    release_header = soup.find('div', id='release-header')
    
    artist_tag = release_header.h1.span
    artist = artist_tag.text
    artist_tag.extract()
    data.append(artist)
#     print(artist)
#     info_dict['Artist'] = artist
    
    title = release_header.h1.text[3:].replace(u'\xa0', u' ')
#     print(title)
    data.append(title)
#     info_dict['title'] = title

    info = release_header.find('div', class_='info_2OHV7')

    format_line = info.find('div', class_='format_3yO6t').text.split(':')[1].split(',')
    if 'LP' in format_line or 'Album' in format_line:
        format_type = 'LP'
    else:
        format_type = 'EP'
    data.append(format_type)
    
    line_items = [div.text.split(':')[1] for div in info.find_all('div', class_='lineitem_2U49R')]
    
    if len(line_items) == 6:
        year = int([int(s) for s in line_items[3].split() if s.isdigit()][0])
        genre = line_items[5]
    else:
        year = int([int(s) for s in line_items[2].split() if s.isdigit()][0])
        genre = line_items[4]
        
    data.append(genre)

    print(line_items)
    data.append(year)

    data.append('')
    
    price = float(soup.find('div', class_='items_3gMeU').find_all('span')[2].text.replace('£', ''))
    data.append(price)
#     info_dict['Format'] = format_line

    print(data)
    
    return data


def get_record_info(link, driver=None):
    if not driver:
        driver = webdriver.Chrome(options=chrome_options)
    driver.get(link)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    soup_info = get_info_from_soup(soup)
    return soup_info


def get_multiple_records_info(links):
    records_info = []
    driver = webdriver.Chrome(options=chrome_options)
    
    for link in links:
        print(link)
        record_info = get_record_info(link, driver=driver)
#         print(record_info)
        records_info.append(record_info)
        
    return records_info
        
# record_info_list = get_multiple_records_info(links)


def append_records_to_file(records, file='test.xlsx'):

    links_excel = openpyxl.load_workbook(file)
    links_worksheet = links_excel.active

    for record in records:
        links_worksheet.append(record)

    # links_worksheet.insert_rows(len(links_worksheet['B']) + 1, amount=1)
    links_excel.save(file)
    links_excel.close()



def append_record_to_file(record, file='test.xlsx'):
    links_excel = openpyxl.load_workbook(file)
    links_worksheet = links_excel.active

    links_worksheet.append(record)

    # links_worksheet.insert_rows(len(links_worksheet['B']) + 1, amount=1)
    links_excel.save(file)
    links_excel.close()



def get_link_from_user():
    link = input('Enter discogs link: ')
    return link


def enter_record(excel_file='records.xlsx'):
    link = get_link_from_user()
    record = get_record_info(link)
    append_record_to_file(record, file=excel_file)

# links_excel = openpyxl.load_workbook('PRE-OWNED STOCK.xlsx')
# links_worksheet = links_excel.active
#
# links_excel
#
# artist_column = [links_worksheet['B'][cell].value for cell in range(len(links_worksheet['B']))][1:]
# title_column = [links_worksheet['C'][cell].value for cell in range(len(links_worksheet['C']))][1:]
# ep_column = [links_worksheet['D'][cell].value for cell in range(len(links_worksheet['D']))][1:]
# genre_column = [links_worksheet['E'][cell].value for cell in range(len(links_worksheet['E']))][1:]
# date_column = [links_worksheet['F'][cell].value for cell in range(len(links_worksheet['F']))][1:]
#
# price_column = [links_worksheet['G'][cell].value for cell in range(len(links_worksheet['G']))][1:]
#
# links_excel.close()
#


def get_links():
    links_excel = openpyxl.load_workbook('PRE-OWNED STOCK.xlsx')
    links_worksheet = links_excel.active

    names_column = links_worksheet['A']
    links_column = links_worksheet['B']
    
    links = []
    for l in range(1, len(links_column)):
        if not links_column[l].value:
            continue
        link = links_column[l].value.replace('\n', '')
        if link[-1] == '/':
            link = link[:-1]
            
        if link not in links: 
            links.append([names_column[l].value, link])
    
    links_excel.close()



if __name__ == '__main__':
    enter_record()